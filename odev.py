# Döviz.com'dan en çok kazandıran 6 hisse senedini çektim. hisse senedi listesi excel dosyasına çekilecek.

import openpyxl
import requests
from bs4 import BeautifulSoup


hisse_url = "https://borsa.doviz.com/"

html_kod = requests.get(hisse_url)



soup = BeautifulSoup(html_kod.content, 'lxml')

baslik=soup.find_all('h3', class_='left')

bilgi_bar = soup.find_all('div', class_='table')[0].find_all('th')

print(baslik[0].get_text())
WorkBook1 = openpyxl.Workbook()
sheet = WorkBook1.active
sheet['A1'] = "Sembol"
sheet['B1'] = "Alış"
sheet['C1'] = "Değişimm"
sheet['D1'] = "Güncelleme"
for i in range(1,7):
    bilgi_bar2 = soup.find_all('div', class_='table')[0].find_all('tr')[i].find_all('td')

    for j in range(0,4):
        print(bilgi_bar[j].get_text()+":"+bilgi_bar2[j].get_text().strip())

    print("----------------------------------------------")

for b in range(2,7):
    bilgi_bar3 = soup.find_all('div', class_='table')[0].find_all('tr')[b].find_all('td')
    for m in range(0,4):
        sheet.cell(row=2, column=m+1).value = bilgi_bar3[b].get_text().strip()
        sheet.cell(row=3, column=m + 1).value = bilgi_bar3[m].get_text().strip()
        sheet.cell(row=4, column=m + 1).value = bilgi_bar3[m].get_text().strip()
        sheet.cell(row=5, column=m + 1).value = bilgi_bar3[m].get_text().strip()
        sheet.cell(row=6, column=m + 1).value = bilgi_bar3[m].get_text().strip()

    # Worknook Kaydet
WorkBook1.save('/Users/ali/Desktop/dovizcom.xlsx')
